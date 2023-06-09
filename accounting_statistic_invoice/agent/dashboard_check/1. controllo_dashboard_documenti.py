#!/usr/bin/python
# -*- coding: utf-8 -*-
###############################################################################
#
# ODOO (ex OpenERP)
# Open Source Management Solution
# Copyright (C) 2001-2015 Micronaet S.r.l. (<https://micronaet.com>)
# Developer: Nicola Riolini @thebrush (<https://it.linkedin.com/in/thebrush>)
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
# See the GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################

import os
import sys
import pdb

# Excel:
import openpyxl
from openpyxl.styles import Border, Alignment, Font, numbers

start = 'statistic.partner.FIA.2023'
partner_db = {}
compare_partner = {}
compare_document = {}
compare = {}

verbose = False
columns = 0
for root, folders, files in os.walk('./data'):
    for filename in files:
        if not filename.startswith(start):
            print('>> Jump %s' % filename)
            continue

        print('Checking %s' % filename)
        file_code = filename[22:-4]
        fullname = os.path.join(root, filename)
        compare[file_code] = {}

        counter = 0
        total = 0.0
        for line in open(fullname, 'r'):
            counter += 1
            if counter == 1:
                if verbose:
                    print('%s: %s. salto intestazione' % (filename, counter))
                continue

            line = line.strip()
            if not line:
                if verbose:
                    print('%s. colonna vuota' % counter)
                continue

            row = line.split('|')
            if not columns:
                columns = len(row)

            if len(row) != columns:
                if verbose:
                    print('%s. colonne differenti' % counter)
                continue

            partner_name = row[1].strip()
            partner_code = row[2].strip()

            month = row[4].strip()
            year = row[5].strip()
            season = row[6].strip()
            period = '%s-%s' % (year, month)

            document = row[7].strip()
            amount = float(row[8].strip().replace(',', '.'))

            if season != '1':
                if verbose:
                    print('%s. stagione non corrente %s' % (counter, season))
                continue

            # Compare all document data:
            if document not in compare[file_code]:
                compare[file_code][document] = 0.0

            compare[file_code][document] += amount

            ''' 
            if partner_code not in compare_partner:
                partner_db[partner_code] = partner_name
                compare_partner[partner_code] = [0.0, 0.0]
                compare_document[partner_code] = {}
            
            if period not in compare_document[partner_code]:
                compare_document[partner_code][period] = {}
                
            if document not in compare_document[partner_code][period]:
                compare_document[partner_code][period][document] = [0.0, 0.0]

                
            compare_partner[partner_code] += amount
            compare_document[partner_code][period][document] += amount
            '''

            total += amount

        print('TOTALI: %s = %s' % (fullname, total))

# Compare file:

# -----------------------------------------------------------------------------
# Style:
# -----------------------------------------------------------------------------
format_euro = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

font = Font(
    name='Verdana', size=10, bold=False, italic=False,
    vertAlign=None, underline='none', strike=False,
    color='FF000000')
font_bold = Font(
    name='Verdana', size=10, bold=True, italic=False,
    vertAlign=None, underline='none', strike=False,
    color='FF000000')

alignment = Alignment(
    horizontal='general', vertical='bottom', text_rotation=0,
    wrap_text=False, shrink_to_fit=False, indent=0)
alignment_center = Alignment(
    horizontal='center', vertical='bottom', text_rotation=0,
    wrap_text=False, shrink_to_fit=False, indent=0)

# -----------------------------------------------------------------------------
#                         Open Excel file:
# -----------------------------------------------------------------------------
filename = './result/documenti_totali.xlsx'
filename = os.path.expanduser(filename)
total_sheet = 'Totali'

# wb = openpyxl.load_workbook(filename=filename)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = total_sheet
# ws = wb.create_sheet(total_sheet)

# Setup columns:
ws.column_dimensions['A'].width = 30
for col_name in 'BCDEFGHI':
    ws.column_dimensions[col_name].width = 25
ws.column_dimensions['L'].width = 40

# -----------------------------------------------------------------------------
# Header:
# -----------------------------------------------------------------------------
header = [
    'Data ora',
    'OC', 'BC', 'FT', 'Totale',
    'Delta OC', 'Delta BC', 'Delta FT', 'Delta Tot.',
    ]
row = 1
col = 0
for title in header:
    col += 1
    cell = ws.cell(row=row, column=col)
    cell.font = font_bold
    cell.alignment = alignment_center
    cell.value = title

# -----------------------------------------------------------------------------
# Data
# -----------------------------------------------------------------------------
previous = False
for file_code in sorted(compare):
    row += 1

    # Read data:
    oc = compare[file_code].get('oo', 0.0)
    bc = compare[file_code].get('bo', 0.0)
    ft = compare[file_code].get('ft', 0.0)
    total = oc + bc + ft

    # Integarte data:
    if not previous:  # Difference:
        previous = [
            oc,
            bc,
            ft,
            total,
        ]
    record = [file_code, oc, bc, ft, total]
    record.extend([
        oc - previous[0],
        bc - previous[1],
        ft - previous[2],
        total - previous[3],
    ])
    # Save this as previous:
    previous = [
        oc,
        bc,
        ft,
        total,
    ]

    col = 0
    for data in record:
        col += 1
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = alignment
        cell.value = data
wb.save(filename)

sys.exit()

f_out = open('./result/documenti_totali.csv', 'w')
f_out.write('Data ora|OC|BC|FT|Totale|Delta OC|Delta BC|Delta FT|Delta Tot.\n')
previous = False
for file_code in sorted(compare):
    oc = compare[file_code].get('oo', 0.0)
    bc = compare[file_code].get('bo', 0.0)
    ft = compare[file_code].get('ft', 0.0)
    total = oc + bc + ft

    # -------------------------------------------------------------------------
    # Difference management:
    # -------------------------------------------------------------------------
    if not previous:
        previous = [
            oc,
            bc,
            ft,
            total,
        ]
    record = [file_code, oc, bc, ft, total]
    record.extend([
        oc - previous[0],
        bc - previous[1],
        ft - previous[2],
        total - previous[3],
    ])
    # Save this as previous:
    previous = [
        oc,
        bc,
        ft,
        total,
    ]
    line = '%s|%s|%s|%s|%s|%s|%s|%s|%s\n' % tuple(record)
    line = line.replace('.', ',')
    f_out.write(line)


'''
    # Create partner compare:
    f_out = open('./result/partner.csv', 'w')
    f_out.write('Codice|Nome|Da|A|Differenza|Note\n')
    for partner_code in compare_partner:
        partner_name = partner_db[partner_code]    
        last, current = compare_partner[partner_code]
        difference = current - last
        if abs(difference) < 0.00001:
            difference = 0.0
        
        if difference < 0:
            error = 'Negativa'
        else:
            error = ''
            
        f_out.write('%s|%s|%s|%s|%s|%s\n' % (
            partner_code, 
            '',  #partner_name, 
            last, current, difference,
            error,
            ))

    # Create partner period document compare:
    f_out = open('./result/document.csv', 'w')
    f_out.write('Codice|Nome|Periodo|Documento|Da|A|Differenza|Note\n')
    for partner_code in compare_document:
        for period in compare_document[partner_code]:
            for document in compare_document[partner_code][period]:
                partner_name = partner_db[partner_code]    
                last, current = compare_document[partner_code][period][document]
                difference = current - last
                if abs(difference) < 0.00001:
                    difference = 0.0
                
                if difference < 0:
                    error = 'Negativa'
                else:
                    error = ''
                    
                f_out.write('%s|%s|%s|%s|%s|%s|%s|%s\n' % (
                    partner_code, 
                    '',  #partner_name, 
                    period, 
                    document,
                    str(last).replace('.', ','), 
                    str(current).replace('.', ','), 
                    str(difference).replace('.', ','), 
                    error,
                    ))
                    
'''
